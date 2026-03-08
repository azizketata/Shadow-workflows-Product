"""Golden Dataset Comparison: Pipeline vs. Human-Annotated Ground Truth.

Compares the automated Meeting Process Twin pipeline outputs against
10 manually annotated Seattle meetings (Whisper medium + human-in-the-loop).

Produces:
  - thesis/figures/golden_shadow_comparison.png
  - thesis/figures/golden_agenda_coverage.png
  - thesis/figures/golden_deviance_comparison.png
  - thesis/figures/golden_meeting_level.png
  - golden_comparison_results.json
"""

import json
import os
import sys
from collections import Counter

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

# --- Configuration ---

GOLDEN_DIR = os.path.join(os.path.dirname(__file__), "Golden-dataset")
MEETINGS_DIR = os.path.join(os.path.dirname(__file__), "meetings")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "thesis", "figures")

# Mapping: golden folder name -> pipeline meeting directory name
GOLDEN_TO_PIPELINE = {
    "Seattle-Council": "seattle_2026-01-20",
    "Seattle-Council-012726": "seattle_2026-01-27",
    "Seattle-Council-020326": "seattle_2026-02-03",
    "Seattle-brief": "seattle_2026-01-26",
    "Seattle-brief-feb": "seattle_2026-02-02",
    "Seattle-finance": "seattle_2026-02-09",
    "Seattle-hous": "seattle_2026-01-28",
    "Seattle-lib": "seattle_2026-01-28",  # same date, different committee
    # No pipeline match:
    # "Seattle-land": "seattle_2026-02-04",  (not in pipeline)
    # "Seattle-transpo": "seattle_2026-01-15",  (not in pipeline)
}


def load_golden(folder_name: str) -> dict:
    """Load golden dataset from Excel."""
    path = os.path.join(GOLDEN_DIR, folder_name, "golden_dataset.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)

    # Events sheet
    ws = wb["Events"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    events = [dict(zip(headers, r)) for r in rows[1:]]

    # Meeting sheet
    ws2 = wb["Meeting"]
    meeting_rows = list(ws2.iter_rows(values_only=True))
    meeting_headers = list(meeting_rows[0])
    meeting_data = dict(zip(meeting_headers, meeting_rows[1])) if len(meeting_rows) > 1 else {}

    wb.close()
    return {"events": events, "meeting": meeting_data, "name": folder_name}


def load_pipeline(meeting_id: str) -> dict | None:
    """Load pipeline outputs for a meeting."""
    meeting_dir = os.path.join(MEETINGS_DIR, meeting_id)
    conf_path = os.path.join(meeting_dir, "conformance.json")
    mapped_path = os.path.join(meeting_dir, "variant_llm", "mapped_events.csv")
    agenda_path = os.path.join(meeting_dir, "agenda.txt")

    if not os.path.exists(conf_path):
        return None

    with open(conf_path) as f:
        conf = json.load(f)

    mapped_df = None
    if os.path.exists(mapped_path):
        mapped_df = pd.read_csv(mapped_path)

    agenda = []
    if os.path.exists(agenda_path):
        with open(agenda_path) as f:
            agenda = [line.strip() for line in f if line.strip()]

    return {"conf": conf, "mapped_df": mapped_df, "agenda": agenda, "meeting_id": meeting_id}


def compute_golden_metrics(golden: dict) -> dict:
    """Compute metrics from golden dataset."""
    events = golden["events"]
    meeting = golden["meeting"]
    total = len(events)

    # Shadow detection
    shadow_count = sum(1 for e in events if e.get("is_shadow") is True)
    formal_count = total - shadow_count
    shadow_pct = shadow_count / total if total > 0 else 0

    # Deviance categories (shadow events only)
    shadow_events = [e for e in events if e.get("is_shadow") is True]
    deviance_cats = Counter()
    for e in shadow_events:
        cat = e.get("deviance_category")
        if cat:
            deviance_cats[cat] += 1
        else:
            deviance_cats["unclassified"] += 1

    # Agenda items covered
    agenda_items_covered = set()
    for e in events:
        match = e.get("agenda_item_match")
        if match and match != "NONE" and not e.get("is_shadow"):
            agenda_items_covered.add(match)

    # Robert's Rules
    roberts = {
        "call_to_order": meeting.get("call_to_order_present", "no") != "no",
        "roll_call": meeting.get("roll_call_present", "no") != "no",
        "adjournment": meeting.get("adjournment_present", "no") != "no",
        "motions_seconded": meeting.get("motions_seconded", "never") not in ("never", "no"),
        "votes_after_motion": meeting.get("votes_after_motion", "never") not in ("never", "no"),
        "public_comment": meeting.get("public_comment_before_vote", "no") not in ("never", "no"),
    }

    return {
        "total_events": total,
        "shadow_count": shadow_count,
        "formal_count": formal_count,
        "shadow_pct": round(shadow_pct, 4),
        "deviance_cats": dict(deviance_cats),
        "agenda_items_covered": sorted(agenda_items_covered),
        "agenda_coverage_count": len(agenda_items_covered),
        "meeting_type": meeting.get("meeting_type", "unknown"),
        "has_formal_agenda": meeting.get("has_formal_agenda", "unknown"),
        "formality": meeting.get("overall_formality", "unknown"),
        "estimated_formal_pct": meeting.get("estimated_formal_pct"),
        "estimated_shadow_pct": meeting.get("estimated_shadow_pct"),
        "roberts_rules": roberts,
        "agenda_order_followed": meeting.get("agenda_order_followed", "unknown"),
    }


def compute_pipeline_metrics(pipeline: dict) -> dict:
    """Extract metrics from pipeline conformance.json."""
    conf = pipeline["conf"]

    ea = conf.get("event_allocation", {})
    fit = conf.get("fitness", {})
    dec = conf.get("declare", {})
    dev = conf.get("deviance", {})
    powl = conf.get("powl", {})

    return {
        "total_events": ea.get("formal", 0) + ea.get("shadow", 0),
        "shadow_count": ea.get("shadow", 0),
        "formal_count": ea.get("formal", 0),
        "shadow_pct": round(ea.get("shadow_pct", 0), 4),
        "dedup_fitness": fit.get("dedup", 0),
        "raw_fitness": fit.get("raw", 0),
        "match_rate": fit.get("match_rate", 0),
        "agenda_coverage_pct": fit.get("agenda_coverage_pct", 0),
        "agenda_coverage_count": fit.get("agenda_coverage", 0),
        "declare_score": dec.get("compliance_score", 0),
        "declare_grade": dec.get("grade", "?"),
        "deviance": {
            "benign": dev.get("benign", 0),
            "innovation": dev.get("innovation", 0),
            "disruption": dev.get("disruption", 0),
            "efficiency": dev.get("efficiency", 0),
        },
        "powl_clusters": powl.get("cluster_count", 0),
    }


def compare_shadow_detection(golden_m: dict, pipeline_m: dict) -> dict:
    """Compare shadow prevalence at the meeting level."""
    return {
        "golden_shadow_pct": golden_m["shadow_pct"],
        "pipeline_shadow_pct": pipeline_m["shadow_pct"],
        "golden_shadow_count": golden_m["shadow_count"],
        "pipeline_shadow_count": pipeline_m["shadow_count"],
        "golden_total": golden_m["total_events"],
        "pipeline_total": pipeline_m["total_events"],
        "agreement_direction": (
            "both_high" if golden_m["shadow_pct"] > 0.5 and pipeline_m["shadow_pct"] > 0.5
            else "both_low" if golden_m["shadow_pct"] <= 0.5 and pipeline_m["shadow_pct"] <= 0.5
            else "disagree"
        ),
    }


def compare_roberts_rules(golden_m: dict, pipeline_m: dict) -> dict:
    """Compare Robert's Rules compliance."""
    golden_roberts = golden_m["roberts_rules"]
    # Pipeline uses Declare; map to comparable checks
    pipeline_declare_grade = pipeline_m["declare_grade"]

    golden_score = sum(golden_roberts.values()) / len(golden_roberts) if golden_roberts else 0

    return {
        "golden_roberts_checks": golden_roberts,
        "golden_roberts_score": round(golden_score, 3),
        "pipeline_declare_grade": pipeline_declare_grade,
        "pipeline_declare_score": pipeline_m["declare_score"],
    }


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    results = []

    print(f"Golden Dataset Comparison: Pipeline vs. Human Annotations")
    print(f"{'='*70}\n")

    for golden_name, pipeline_id in GOLDEN_TO_PIPELINE.items():
        golden = load_golden(golden_name)
        pipeline = load_pipeline(pipeline_id)

        if pipeline is None:
            print(f"  SKIP {golden_name} -> {pipeline_id} (no pipeline data)")
            continue

        golden_m = compute_golden_metrics(golden)
        pipeline_m = compute_pipeline_metrics(pipeline)
        shadow_cmp = compare_shadow_detection(golden_m, pipeline_m)
        roberts_cmp = compare_roberts_rules(golden_m, pipeline_m)

        result = {
            "golden_name": golden_name,
            "pipeline_id": pipeline_id,
            "golden": golden_m,
            "pipeline": pipeline_m,
            "shadow_comparison": shadow_cmp,
            "roberts_comparison": roberts_cmp,
        }
        results.append(result)

        # Print summary
        print(f"  {golden_name:25s} -> {pipeline_id}")
        print(f"    Events:     golden={golden_m['total_events']:5d}   pipeline={pipeline_m['total_events']:3d}   "
              f"(ratio: {golden_m['total_events']/max(pipeline_m['total_events'],1):.0f}x)")
        print(f"    Shadow%:    golden={golden_m['shadow_pct']*100:.1f}%    pipeline={pipeline_m['shadow_pct']*100:.1f}%   "
              f"({shadow_cmp['agreement_direction']})")
        print(f"    Deviance:   golden={golden_m['deviance_cats']}")
        print(f"                pipeline={pipeline_m['deviance']}")
        print(f"    Roberts:    golden_score={roberts_cmp['golden_roberts_score']:.2f}   "
              f"pipeline_declare={roberts_cmp['pipeline_declare_grade']}")
        print()

    if not results:
        print("No matched meetings found!")
        sys.exit(1)

    # Aggregate statistics
    print(f"\n{'='*70}")
    print(f"AGGREGATE STATISTICS ({len(results)} matched meetings)")
    print(f"{'='*70}\n")

    golden_shadows = [r["shadow_comparison"]["golden_shadow_pct"] for r in results]
    pipeline_shadows = [r["shadow_comparison"]["pipeline_shadow_pct"] for r in results]
    print(f"Shadow Prevalence:")
    print(f"  Golden  mean: {np.mean(golden_shadows)*100:.1f}%  median: {np.median(golden_shadows)*100:.1f}%")
    print(f"  Pipeline mean: {np.mean(pipeline_shadows)*100:.1f}%  median: {np.median(pipeline_shadows)*100:.1f}%")

    agreements = [r["shadow_comparison"]["agreement_direction"] for r in results]
    print(f"  Direction agreement: {Counter(agreements)}")

    # Correlation
    corr = np.corrcoef(golden_shadows, pipeline_shadows)[0, 1]
    print(f"  Pearson correlation: {corr:.3f}")

    # Deviance aggregates
    golden_dev_total = Counter()
    pipeline_dev_total = Counter()
    for r in results:
        for k, v in r["golden"]["deviance_cats"].items():
            golden_dev_total[k] += v
        for k, v in r["pipeline"]["deviance"].items():
            pipeline_dev_total[k] += v

    print(f"\nDeviance Distribution (aggregated):")
    print(f"  Golden:   {dict(golden_dev_total)}")
    print(f"  Pipeline: {dict(pipeline_dev_total)}")

    # Roberts Rules
    golden_roberts_scores = [r["roberts_comparison"]["golden_roberts_score"] for r in results]
    pipeline_declare_scores = [r["roberts_comparison"]["pipeline_declare_score"] for r in results]
    print(f"\nRobert's Rules / Declare:")
    print(f"  Golden Roberts mean:  {np.mean(golden_roberts_scores):.3f}")
    print(f"  Pipeline Declare mean: {np.mean(pipeline_declare_scores):.3f}")

    # Save results
    results_path = os.path.join(os.path.dirname(__file__), "golden_comparison_results.json")
    with open(results_path, "w") as f:
        json.dump(results, f, indent=2, default=str)
    print(f"\nResults saved to {results_path}")

    # Generate figures
    generate_figures(results)


def generate_figures(results: list[dict]):
    """Generate all golden comparison figures."""
    names = [r["golden_name"].replace("Seattle-", "S-") for r in results]
    golden_shadow = [r["shadow_comparison"]["golden_shadow_pct"] * 100 for r in results]
    pipeline_shadow = [r["shadow_comparison"]["pipeline_shadow_pct"] * 100 for r in results]

    # --- Figure 1: Shadow Prevalence Comparison ---
    fig, axes = plt.subplots(1, 2, figsize=(12, 5))

    # Panel a: Side-by-side bars
    ax = axes[0]
    x = np.arange(len(names))
    w = 0.35
    ax.bar(x - w/2, golden_shadow, w, label="Golden (human)", color="#1976D2", alpha=0.8)
    ax.bar(x + w/2, pipeline_shadow, w, label="Pipeline (auto)", color="#E53935", alpha=0.8)
    ax.set_xticks(x)
    ax.set_xticklabels(names, rotation=45, ha="right", fontsize=8)
    ax.set_ylabel("Shadow Activity (%)")
    ax.set_title("(a) Per-Meeting Shadow Prevalence", fontweight="bold")
    ax.legend(fontsize=9)
    ax.grid(axis="y", alpha=0.3)

    # Panel b: Scatter correlation
    ax = axes[1]
    ax.scatter(golden_shadow, pipeline_shadow, s=80, c="#43A047", edgecolors="black", linewidth=0.5, zorder=3)
    for i, name in enumerate(names):
        ax.annotate(name, (golden_shadow[i], pipeline_shadow[i]),
                    textcoords="offset points", xytext=(5, 5), fontsize=7)
    # Identity line
    lim = max(max(golden_shadow), max(pipeline_shadow)) + 5
    ax.plot([0, lim], [0, lim], "k--", alpha=0.3, linewidth=1)
    corr = np.corrcoef(golden_shadow, pipeline_shadow)[0, 1]
    ax.set_xlabel("Golden Shadow %")
    ax.set_ylabel("Pipeline Shadow %")
    ax.set_title(f"(b) Correlation (r = {corr:.3f})", fontweight="bold")
    ax.grid(True, alpha=0.3)

    fig.suptitle("Ground-Truth Validation: Shadow Prevalence", fontsize=13, fontweight="bold", y=1.02)
    fig.tight_layout()
    path = os.path.join(OUTPUT_DIR, "golden_shadow_comparison.png")
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)
    print(f"  Figure saved: {path}")

    # --- Figure 2: Deviance Category Comparison ---
    fig, axes = plt.subplots(1, 2, figsize=(11, 5))

    # Golden deviance
    ax = axes[0]
    golden_dev_total = Counter()
    for r in results:
        for k, v in r["golden"]["deviance_cats"].items():
            golden_dev_total[k] += v
    cats_g = dict(golden_dev_total)
    total_g = sum(cats_g.values())
    labels_g = [f"{k}\n({v/total_g*100:.1f}%)" for k, v in cats_g.items()]
    colors_g = {"benign": "#64B5F6", "innovation": "#81C784", "disruption": "#E57373",
                "efficiency": "#FFD54F", "unclassified": "#BDBDBD"}
    ax.pie(cats_g.values(), labels=labels_g,
           colors=[colors_g.get(k, "#BDBDBD") for k in cats_g.keys()],
           startangle=90, textprops={"fontsize": 9})
    ax.set_title(f"(a) Golden Dataset (n={total_g})", fontweight="bold")

    # Pipeline deviance
    ax = axes[1]
    pipeline_dev_total = Counter()
    for r in results:
        for k, v in r["pipeline"]["deviance"].items():
            pipeline_dev_total[k] += v
    cats_p = {k: v for k, v in pipeline_dev_total.items() if v > 0}
    total_p = sum(cats_p.values())
    if total_p > 0:
        labels_p = [f"{k}\n({v/total_p*100:.1f}%)" for k, v in cats_p.items()]
        ax.pie(cats_p.values(), labels=labels_p,
               colors=[colors_g.get(k, "#BDBDBD") for k in cats_p.keys()],
               startangle=90, textprops={"fontsize": 9})
    ax.set_title(f"(b) Pipeline (n={total_p})", fontweight="bold")

    fig.suptitle("Deviance Classification: Golden vs. Pipeline", fontsize=13, fontweight="bold", y=1.02)
    fig.tight_layout()
    path = os.path.join(OUTPUT_DIR, "golden_deviance_comparison.png")
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)
    print(f"  Figure saved: {path}")

    # --- Figure 3: Meeting-Level Validation Summary ---
    fig, axes = plt.subplots(1, 3, figsize=(14, 4.5))

    # Panel a: Event count comparison (log scale)
    ax = axes[0]
    golden_counts = [r["golden"]["total_events"] for r in results]
    pipeline_counts = [r["pipeline"]["total_events"] for r in results]
    ax.bar(x - w/2, golden_counts, w, label="Golden", color="#1976D2", alpha=0.8)
    ax.bar(x + w/2, pipeline_counts, w, label="Pipeline", color="#E53935", alpha=0.8)
    ax.set_xticks(x)
    ax.set_xticklabels(names, rotation=45, ha="right", fontsize=7)
    ax.set_ylabel("Event Count")
    ax.set_yscale("log")
    ax.set_title("(a) Event Granularity", fontweight="bold")
    ax.legend(fontsize=8)
    ax.grid(axis="y", alpha=0.3)

    # Panel b: Formality assessment
    ax = axes[1]
    golden_formal_pct = [(1 - r["shadow_comparison"]["golden_shadow_pct"]) * 100 for r in results]
    pipeline_formal_pct = [(1 - r["shadow_comparison"]["pipeline_shadow_pct"]) * 100 for r in results]
    ax.bar(x - w/2, golden_formal_pct, w, label="Golden", color="#1976D2", alpha=0.8)
    ax.bar(x + w/2, pipeline_formal_pct, w, label="Pipeline", color="#E53935", alpha=0.8)
    ax.set_xticks(x)
    ax.set_xticklabels(names, rotation=45, ha="right", fontsize=7)
    ax.set_ylabel("Formal Activity (%)")
    ax.set_title("(b) Formal Activity Rate", fontweight="bold")
    ax.legend(fontsize=8)
    ax.grid(axis="y", alpha=0.3)

    # Panel c: Roberts Rules comparison
    ax = axes[2]
    golden_roberts = [r["roberts_comparison"]["golden_roberts_score"] * 100 for r in results]
    pipeline_declare = [r["roberts_comparison"]["pipeline_declare_score"] for r in results]
    ax.bar(x - w/2, golden_roberts, w, label="Golden (Roberts)", color="#1976D2", alpha=0.8)
    ax.bar(x + w/2, pipeline_declare, w, label="Pipeline (Declare)", color="#E53935", alpha=0.8)
    ax.set_xticks(x)
    ax.set_xticklabels(names, rotation=45, ha="right", fontsize=7)
    ax.set_ylabel("Compliance Score (%)")
    ax.set_title("(c) Procedural Compliance", fontweight="bold")
    ax.legend(fontsize=8)
    ax.grid(axis="y", alpha=0.3)

    fig.suptitle("Meeting-Level Validation: Pipeline vs. Ground Truth (8 Seattle Meetings)",
                 fontsize=13, fontweight="bold", y=1.02)
    fig.tight_layout()
    path = os.path.join(OUTPUT_DIR, "golden_meeting_level.png")
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)
    print(f"  Figure saved: {path}")


if __name__ == "__main__":
    main()
